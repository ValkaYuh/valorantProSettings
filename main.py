import multiprocessing

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from selenium import webdriver
from selenium.common import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading


# helper function for the update_or_add_entry func
def find_next_empty_row(sheet, col=1):
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=col).value is None:
            return row
    return sheet.max_row + 1


# updating entry if it already exists, adds new entry in the first empty row if it doesn't
def update_or_add_entry(file_path, new_entry):
    workbook = load_workbook(file_path, data_only=False)
    sheet = workbook.active

    # look for an existing entry by name
    for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
        if row[0].value == new_entry[0]:
            update_row = row[0].row
            break
    else:
        update_row = find_next_empty_row(sheet)

    for col, value in enumerate(new_entry, start=1):
        sheet.cell(row=update_row, column=col, value=value)

    workbook.save(file_path)


# helper function for scraping
def wait_for_element(driver, xpath, attr='text', timeout=0.1):
    element = WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.XPATH, xpath))
    )
    return element.text if attr == 'text' else getattr(element, attr)


# helper function for scraping
def wait_for_element_any(driver, xpaths, attr='accessible_name', timeout=0.1, default="unknown"):
    for xpath in xpaths:
        try:
            element = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            return getattr(element, attr)
        except TimeoutException:
            continue
    return default


# scraping all the data we need
def get_info(url):
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    driver = webdriver.Chrome(options=chrome_options)
    try:
        driver.get(url)

        name_value = wait_for_element(driver, "//div[@class='name']/h1", 'text')
        BFI_value = wait_for_element_any(
            driver,
            ["//tr[@class='format-select field-dyac']/td",
             "//tr[@class='format-select field-dyac unknown']/td"],
            'accessible_name', 0.1, "unknown"
        )
        eDPI_value = wait_for_element(driver, "//tr[@class='format-number field-edpi']/td", 'text')
        mousepad_value = wait_for_element(
            driver,
            "//div[contains(@class, 'cta-box') and ./div[@class='cta-box__tag cta-box__tag--top-right' and text()='Mousepad']]//h4/a",
            'accessible_name'
        )
        outline_value = wait_for_element_any(
            driver,
            ["//tr[@class='format-select field-enemyhighlightcolor']/td",
             "//tr[@class='format-select field-enemyhighlightcolor unknown']/td"],
            'accessible_name', 0.1, "unknown"
        )

        # Format outline_value
        if "Yellow" in outline_value:
            outline_value = "yellow"
        elif "Purple" in outline_value:
            outline_value = "purple"
        elif "Red" in outline_value:
            outline_value = "red"
        elif "Unknown" in outline_value:
            outline_value = "unknown"

        # Format BFI_value
        if "Premium" in BFI_value or "High" in BFI_value:
            BFI_value = "ON"
        elif "Off" in BFI_value:
            BFI_value = "OFF"
        elif "Unknown" in BFI_value:
            BFI_value = "unknown"

        eDPI_value = float(eDPI_value.strip("'"))

        return [name_value, eDPI_value, mousepad_value, outline_value, BFI_value]
    finally:
        driver.quit()


lock = threading.Lock()  # prevent simultaneous file writes in threads


def update_whole_list(file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active
    count = 0

    # collect player names from the first column (starting from row 2)
    names = []
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if not row[0]:
            break
        names.append(row[0])

    def process_player(name):
        temp_url = f"https://prosettings.net/players/{name.lower()}/"
        temp_result = get_info(temp_url)
        with lock:
            update_or_add_entry(file_path, temp_result)
        return name

    # leaving one core free just in case (there were no issues when using all cores from my testing)
    with ThreadPoolExecutor(max_workers=multiprocessing.cpu_count() - 1) as executor:
        future_to_name = {executor.submit(process_player, name): name for name in names}
        for future in as_completed(future_to_name):
            name = future_to_name[future]
            try:
                future.result()  # Reraise any exceptions.
                count += 1
                print(f"Updated: {name}")
            except Exception as e:
                print(f"Error updating {name}: {e}")

    return count


# generate a template xlsx sheet, useful for when the script is ran for the first time
def generate_template(file_path="pro players.xlsx"):
    wb = Workbook()
    ws = wb.active

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    black_font = Font(color="000000", bold=True)

    headers = ["NAME", "eDPI", "MOUSEPAD", "OUTLINE", "BFI"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = yellow_fill
        cell.font = black_font

    ws["F5"] = "avg eDPI:"
    ws["F5"].fill = yellow_fill
    ws["F5"].font = black_font

    ws["G5"] = "=IFERROR(AVERAGE(B2:INDEX(B:B,MATCH(1E+306,B:B))), 0)"
    ws["G5"].fill = yellow_fill
    ws["G5"].font = black_font

    ws["H2"] = "BFI:"
    ws["H2"].fill = yellow_fill
    ws["H2"].font = black_font

    ws["I2"] = (
        "=CONCATENATE(\"ON: \", COUNTIF(E:E, \"ON\"), \", OFF: \", COUNTIF(E:E, \"OFF\"), "
        "\" , ON%: \", TEXT(IF(COUNTIF(E:E, \"ON\")+COUNTIF(E:E, \"OFF\")>0, "
        "COUNTIF(E:E, \"ON\")/(COUNTIF(E:E, \"ON\")+COUNTIF(E:E, \"OFF\")), 0), \"0.00%\"))"
    )
    ws["I2"].fill = yellow_fill
    ws["I2"].font = black_font

    wb.save(file_path)
    print(f"Template saved as {file_path}")


# all commands are entered here
def main():
    file_path = "pro players.xlsx"
    user_input = input('please insert the link of the player that you want to import, or any command (use "help" for a list of commands): ').strip().lower()

    if user_input == "update":
        count = update_whole_list(file_path)
        print(f"{count} entries have been updated!")
    elif user_input == "generate":
        generate_template(file_path)
        print("Template has been generated in the script's directory.")
    elif user_input == "help":
        print(" - Pasting any player's link from prosettings.net will add a new entry in the first empty row.\n",
              '- "generate": Generates the base template into which you can start importing valorant players.\n',
              '- "update": Updates all existing entries.\n',
              '- "exit": Closes the script.\n')
    elif user_input == "exit":
        exit()
    else:
        result = get_info(user_input)
        update_or_add_entry(file_path, result)


# just loop main until we get the "exit" command
if __name__ == "__main__":
    while True:
        main()
